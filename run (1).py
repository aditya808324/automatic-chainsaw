#!/usr/bin/env python3
# run.py — Salon Bot v3 (flat single-file version)
# All modules merged: database, slots, sheets, payments, bot, backend

import asyncio, hashlib, hmac, json, logging, os, sqlite3, sys
import threading, uuid
from contextlib import contextmanager
from datetime import datetime, timedelta, date as date_type
from typing import Optional
from urllib.parse import parse_qs, unquote

from dotenv import load_dotenv
load_dotenv()

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(name)s | %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────
BOT_TOKEN       = os.getenv("BOT_TOKEN",       "YOUR_BOT_TOKEN")
ADMIN_CHAT_ID   = int(os.getenv("ADMIN_CHAT_ID","0"))
MINI_APP_URL    = os.getenv("MINI_APP_URL",    "https://your-vercel-app.vercel.app")
DB_PATH         = os.getenv("DB_PATH",         "salon.db")
SHEET_ID        = os.getenv("SHEET_ID",        "")
SHEET_NAME      = os.getenv("SHEET_NAME",      "Bookings")
GOOGLE_CREDS    = os.getenv("GOOGLE_CREDS",    "google_credentials.json")
RAZORPAY_KEY    = os.getenv("RAZORPAY_KEY",    "")
RAZORPAY_SECRET = os.getenv("RAZORPAY_SECRET", "")
ADVANCE_AMOUNT  = int(os.getenv("ADVANCE_AMOUNT","100"))
PORT            = int(os.getenv("PORT",         "8000"))

# ── Database ──────────────────────────────────────────────────────────
@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    try:
        yield conn; conn.commit()
    except:
        conn.rollback(); raise
    finally:
        conn.close()

def init_db():
    with get_db() as db:
        db.execute("""CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT, telegram_id INTEGER UNIQUE,
            name TEXT NOT NULL, phone TEXT, visit_count INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')), last_seen TEXT DEFAULT (datetime('now')))""")
        db.execute("""CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
            duration INTEGER NOT NULL DEFAULT 30, price INTEGER NOT NULL DEFAULT 0, active INTEGER DEFAULT 1)""")
        db.execute("""CREATE TABLE IF NOT EXISTS staff (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, active INTEGER DEFAULT 1)""")
        db.execute("""CREATE TABLE IF NOT EXISTS bookings (
            id TEXT PRIMARY KEY, telegram_id INTEGER, client_name TEXT NOT NULL, phone TEXT,
            service TEXT NOT NULL, staff TEXT NOT NULL, date TEXT NOT NULL, slot TEXT NOT NULL,
            duration INTEGER DEFAULT 30, total_price INTEGER DEFAULT 0, advance_amount INTEGER DEFAULT 0,
            payment_status TEXT DEFAULT 'pending', payment_id TEXT, conflict_flag INTEGER DEFAULT 0,
            status TEXT DEFAULT 'confirmed', reminder_sent INTEGER DEFAULT 0, notes TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now')))""")
        db.execute("""CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL)""")
        for name,dur,price in [("Haircut",30,200),("Beard Trim",20,100),("Facial",60,500),("Hair Spa",90,700)]:
            db.execute("INSERT OR IGNORE INTO services (name,duration,price) VALUES (?,?,?)",(name,dur,price))
        for name in ["Priya","Rahul","Neha"]:
            db.execute("INSERT OR IGNORE INTO staff (name) VALUES (?)",(name,))
        for k,v in {"salon_name":"Shringar Beauty Studio","salon_address":"Connaught Place, New Delhi",
            "salon_phone":"+91 98765 43210","open_time":"09:00","close_time":"20:00",
            "slot_interval":"30","currency":"₹","working_days":"1,2,3,4,5,6"}.items():
            db.execute("INSERT OR IGNORE INTO settings (key,value) VALUES (?,?)",(k,v))
    logger.info(f"[DB] Initialized: {DB_PATH}")

def get_setting(key,default=""):
    with get_db() as db:
        row = db.execute("SELECT value FROM settings WHERE key=?",(key,)).fetchone()
        return row["value"] if row else default

def set_setting(key,value):
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)",(key,value))

def get_services():
    with get_db() as db:
        return db.execute("SELECT * FROM services WHERE active=1 ORDER BY id").fetchall()

def get_staff_list():
    with get_db() as db:
        return db.execute("SELECT * FROM staff WHERE active=1 ORDER BY id").fetchall()

def get_booked_slots(date_str,staff_name=None):
    with get_db() as db:
        if staff_name and staff_name!="Any Available":
            rows=db.execute("SELECT slot,duration FROM bookings WHERE date=? AND staff=? AND status!='cancelled'",(date_str,staff_name)).fetchall()
        else:
            rows=db.execute("SELECT slot,duration FROM bookings WHERE date=? AND status!='cancelled'",(date_str,)).fetchall()
        return [(r["slot"],r["duration"]) for r in rows]

def check_conflict(date_str,slot,staff_name):
    return any(s==slot for s,_ in get_booked_slots(date_str,staff_name))

def save_booking(data):
    ref="SHR-"+uuid.uuid4().hex[:6].upper()
    conflict=check_conflict(data["date"],data["slot"],data["staff"])
    with get_db() as db:
        db.execute("INSERT INTO bookings (id,telegram_id,client_name,phone,service,staff,date,slot,duration,total_price,advance_amount,payment_status,conflict_flag,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (ref,data.get("telegram_id"),data["client_name"],data.get("phone"),data["service"],data["staff"],
             data["date"],data["slot"],data.get("duration",30),data.get("total_price",0),
             data.get("advance_amount",0),data.get("payment_status","pending"),1 if conflict else 0,data.get("notes","")))
        db.execute("INSERT INTO clients (telegram_id,name,phone,visit_count) VALUES (?,?,?,1) ON CONFLICT(telegram_id) DO UPDATE SET name=excluded.name,phone=excluded.phone,visit_count=visit_count+1,last_seen=datetime('now')",
            (data.get("telegram_id"),data["client_name"],data.get("phone")))
    return ref,conflict

def get_todays_bookings():
    today=date_type.today().isoformat()
    with get_db() as db:
        return db.execute("SELECT * FROM bookings WHERE date=? AND status!='cancelled' ORDER BY slot",(today,)).fetchall()

def get_recent_clients(limit=10):
    with get_db() as db:
        return db.execute("SELECT * FROM clients ORDER BY last_seen DESC LIMIT ?",(limit,)).fetchall()

def get_revenue_today():
    today=date_type.today().isoformat()
    with get_db() as db:
        row=db.execute("SELECT COUNT(*) as count,COALESCE(SUM(advance_amount),0) as revenue FROM bookings WHERE date=? AND status!='cancelled'",(today,)).fetchone()
    return {"count":row["count"],"revenue":row["revenue"]}

def update_payment_db(booking_id,payment_id,status):
    with get_db() as db:
        db.execute("UPDATE bookings SET payment_id=?,payment_status=? WHERE id=?",(payment_id,status,booking_id))

# ── Slots ─────────────────────────────────────────────────────────────
def generate_all_slots(open_time,close_time,interval_mins):
    slots,fmt=[],"%H:%M"
    cur=datetime.strptime(open_time,fmt); end=datetime.strptime(close_time,fmt)
    while cur<end:
        slots.append(cur.strftime(fmt)); cur+=timedelta(minutes=interval_mins)
    return slots

def get_slots_for_date(date_str,staff_name=None):
    open_t=get_setting("open_time","09:00"); close_t=get_setting("close_time","20:00")
    interval=int(get_setting("slot_interval","30"))
    working=[int(d) for d in get_setting("working_days","1,2,3,4,5,6").split(",")]
    date_obj=datetime.strptime(date_str,"%Y-%m-%d")
    if date_obj.weekday() not in working:
        return {"available":[],"booked":[],"all":[],"closed":True}
    all_slots=generate_all_slots(open_t,close_t,interval)
    booked_set={s for s,_ in get_booked_slots(date_str,staff_name)}
    now=datetime.now()
    if date_str==now.strftime("%Y-%m-%d"):
        all_slots=[s for s in all_slots if datetime.strptime(s,"%H:%M").replace(year=now.year,month=now.month,day=now.day)>now+timedelta(minutes=30)]
    return {"available":[s for s in all_slots if s not in booked_set],"booked":[s for s in all_slots if s in booked_set],"all":all_slots,"closed":False}

# ── Google Sheets ─────────────────────────────────────────────────────
def append_to_sheet(booking):
    if not SHEET_ID or not os.path.exists(GOOGLE_CREDS): return False
    try:
        import gspread; from google.oauth2.service_account import Credentials
        creds=Credentials.from_service_account_file(GOOGLE_CREDS,scopes=["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive"])
        sheet=gspread.authorize(creds).open_by_key(SHEET_ID).worksheet(SHEET_NAME)
        if not sheet.row_values(1):
            sheet.append_row(["Timestamp","ID","Client","Phone","Service","Staff","Date","Slot","Duration","Total","Advance","Payment","Conflict"])
        sheet.append_row([datetime.now().strftime("%Y-%m-%d %H:%M:%S"),booking.get("id",""),booking.get("client_name",""),
            booking.get("phone",""),booking.get("service",""),booking.get("staff",""),booking.get("date",""),
            booking.get("slot",""),booking.get("duration",""),f"₹{booking.get('total_price',0)}",
            f"₹{booking.get('advance_amount',0)}",booking.get("payment_status","pending"),
            "⚠️ CONFLICT" if booking.get("conflict_flag") else "OK"])
        return True
    except Exception as e:
        logger.warning(f"[SHEETS] Failed: {e}"); return False

# ── FastAPI ───────────────────────────────────────────────────────────
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn, httpx

api = FastAPI(title="Salon Bot v3",version="3.0")
api.add_middleware(CORSMiddleware,allow_origins=["*"],allow_methods=["*"],allow_headers=["*"])

@api.on_event("startup")
async def startup(): init_db(); logger.info("[API] Started")

def verify_init_data(init_data):
    try:
        parsed=dict(parse_qs(unquote(init_data),keep_blank_values=True))
        flat={k:v[0] for k,v in parsed.items()}; hash_=flat.pop("hash",None)
        if not hash_: return None
        data_check="\n".join(f"{k}={v}" for k,v in sorted(flat.items()))
        secret=hmac.new(b"WebAppData",BOT_TOKEN.encode(),hashlib.sha256).digest()
        expected=hmac.new(secret,data_check.encode(),hashlib.sha256).hexdigest()
        return json.loads(flat.get("user","{}")) if hmac.compare_digest(expected,hash_) else None
    except: return None

class BookingRequest(BaseModel):
    init_data:str; client_name:str; phone:str; service:str; staff:str
    date:str; slot:str; duration:int; total_price:int; notes:Optional[str]=""

class PaymentVerifyRequest(BaseModel):
    booking_id:str; order_id:str; payment_id:str; signature:str

@api.get("/health")
async def health(): return {"status":"ok","version":"3.0"}

@api.get("/api/config")
async def api_config():
    return {"salon_name":get_setting("salon_name"),"salon_address":get_setting("salon_address"),
            "salon_phone":get_setting("salon_phone"),"open_time":get_setting("open_time"),
            "close_time":get_setting("close_time"),"currency":get_setting("currency"),
            "advance_amount":ADVANCE_AMOUNT,"razorpay_key":RAZORPAY_KEY}

@api.get("/api/services")
async def api_services():
    return [{"id":r["id"],"name":r["name"],"duration":r["duration"],"price":r["price"]} for r in get_services()]

@api.get("/api/staff")
async def api_staff_route():
    return [{"id":r["id"],"name":r["name"]} for r in get_staff_list()]

@api.get("/api/slots")
async def api_slots(date:str,staff:str="Any Available"):
    try: datetime.strptime(date,"%Y-%m-%d")
    except: raise HTTPException(400,"Invalid date")
    return get_slots_for_date(date,staff)

@api.post("/api/book")
async def create_booking(req:BookingRequest):
    user=verify_init_data(req.init_data) or {"id":0,"first_name":req.client_name}
    data={"telegram_id":user.get("id",0),"client_name":req.client_name,"phone":req.phone,
          "service":req.service,"staff":req.staff,"date":req.date,"slot":req.slot,
          "duration":req.duration,"total_price":req.total_price,
          "advance_amount":ADVANCE_AMOUNT,"payment_status":"pending","notes":req.notes}
    ref,conflict=save_booking(data); data["id"]=ref
    try: append_to_sheet(data)
    except: pass
    try: await _admin_alert(data,conflict)
    except: pass
    logger.info(f"[API] Booking {ref} | conflict={conflict}")
    return {"booking_id":ref,"conflict":conflict,"order_id":None,"amount":ADVANCE_AMOUNT,"razorpay_key":RAZORPAY_KEY}

@api.post("/api/payment/verify")
async def verify_payment(req:PaymentVerifyRequest):
    update_payment_db(req.booking_id,req.payment_id,"paid")
    return {"status":"success"}

async def _admin_alert(booking,conflict):
    cur=get_setting("currency","₹")
    banner="\n⚠️ *DOUBLE BOOKING ALERT!*\n" if conflict else ""
    text=(f"🔔 *New Booking*{banner}\n━━━━━━━━━━━━━━━━━━━━\n\n"
          f"👤 *Client:* {booking['client_name']}\n📞 *Phone:* {booking.get('phone','—')}\n\n"
          f"🪄 *Service:* {booking['service']}\n👩‍🎨 *Staff:* {booking['staff']}\n"
          f"📅 *Date:* {booking['date']}\n⏰ *Slot:* {booking['slot']}\n"
          f"💰 *Total:* {cur}{booking.get('total_price',0)}\n\n🔖 *ID:* `{booking['id']}`")
    async with httpx.AsyncClient() as c:
        await c.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id":ADMIN_CHAT_ID,"text":text,"parse_mode":"Markdown"},timeout=10)

# ── Telegram Bot ──────────────────────────────────────────────────────
from telegram import Update,InlineKeyboardButton,InlineKeyboardMarkup,WebAppInfo
from telegram.ext import Application,CommandHandler,CallbackQueryHandler,ContextTypes

def is_admin(uid): return uid==ADMIN_CHAT_ID
def book_kb(): return InlineKeyboardMarkup([[InlineKeyboardButton("✦ Book Appointment",web_app=WebAppInfo(url=MINI_APP_URL))]])

async def start(update:Update,context:ContextTypes.DEFAULT_TYPE):
    salon=get_setting("salon_name","Shringar Studio")
    await update.message.reply_text(
        f"🌸 *Namaste {update.effective_user.first_name}!*\n\nWelcome to *{salon}*\n\nTap below to book 👇",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✦ Book Appointment",web_app=WebAppInfo(url=MINI_APP_URL))],
            [InlineKeyboardButton("📋 My Bookings",callback_data="my_bookings"),
             InlineKeyboardButton("ℹ️ About",callback_data="about")],
        ]))

async def admin_cmd(update:Update,context:ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): await update.message.reply_text("⛔ Admin only."); return
    r=get_revenue_today(); cur=get_setting("currency","₹")
    await update.message.reply_text(
        f"⚙️ *Admin Panel*\n\n📅 Today: *{r['count']} bookings*\n💰 Revenue: *{cur}{r['revenue']:,}*",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("📅 Today",callback_data="admin_today"),
             InlineKeyboardButton("💰 Revenue",callback_data="admin_revenue")],
            [InlineKeyboardButton("👥 Clients",callback_data="admin_clients"),
             InlineKeyboardButton("📤 Export",callback_data="admin_export")],
        ]))

async def status_cmd(update:Update,context:ContextTypes.DEFAULT_TYPE):
    r=get_revenue_today(); cur=get_setting("currency","₹")
    with get_db() as db:
        total=db.execute("SELECT COUNT(*) as c FROM bookings").fetchone()["c"]
        clients=db.execute("SELECT COUNT(*) as c FROM clients").fetchone()["c"]
    await update.message.reply_text(
        f"🔧 *Bot v3.0*\n\n✅ Running\n📅 Today: {r['count']}\n💰 {cur}{r['revenue']:,}\n📋 Total: {total} | 👥 {clients}",
        parse_mode="Markdown")

async def setname_cmd(update:Update,context:ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    val=" ".join(context.args)
    if not val: await update.message.reply_text("Usage: /setname My Salon"); return
    set_setting("salon_name",val); await update.message.reply_text(f"✅ *{val}*",parse_mode="Markdown")

async def sethours_cmd(update:Update,context:ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    if len(context.args)!=2: await update.message.reply_text("Usage: /sethours 09:00 20:00"); return
    set_setting("open_time",context.args[0]); set_setting("close_time",context.args[1])
    await update.message.reply_text(f"✅ *{context.args[0]}* – *{context.args[1]}*",parse_mode="Markdown")

async def addservice_cmd(update:Update,context:ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    parts=" ".join(context.args).split("|")
    if len(parts)!=3: await update.message.reply_text("Usage: /addservice Haircut|30|200"); return
    with get_db() as db:
        db.execute("INSERT INTO services (name,duration,price) VALUES (?,?,?)",(parts[0].strip(),int(parts[1].strip()),int(parts[2].strip())))
    await update.message.reply_text(f"✅ Added: *{parts[0].strip()}*",parse_mode="Markdown")

async def addstaff_cmd(update:Update,context:ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    name=" ".join(context.args).strip()
    if not name: await update.message.reply_text("Usage: /addstaff Priya"); return
    with get_db() as db: db.execute("INSERT INTO staff (name) VALUES (?)",(name,))
    await update.message.reply_text(f"✅ Added: *{name}*",parse_mode="Markdown")

async def handle_callback(update:Update,context:ContextTypes.DEFAULT_TYPE):
    q=update.callback_query; await q.answer()
    if q.data=="admin_today": await _show_today(q.message,edit=True)
    elif q.data=="admin_revenue": await _show_revenue(q.message,edit=True)
    elif q.data=="admin_clients": await _show_clients(q.message,edit=True)
    elif q.data=="admin_export": await _send_export(q.message,context)
    elif q.data=="my_bookings": await _show_my_bookings(update,context)
    elif q.data=="about": await _show_about(q.message,edit=True)

async def _show_today(msg,edit=False):
    bookings=get_todays_bookings(); cur=get_setting("currency","₹")
    today=date_type.today().strftime("%A, %d %B %Y")
    if not bookings: text=f"📅 *{today}*\n\nNo appointments today 🎉"
    else:
        lines=[f"📅 *{today}* — {len(bookings)} appointments\n"]
        for b in bookings:
            lines.append(f"{'✅' if b['payment_status']=='paid' else '⏳'} *{b['slot']}* {b['service']}{'⚠️' if b['conflict_flag'] else ''}\n   👤 {b['client_name']} 📞 {b['phone'] or '—'}\n   👩‍🎨 {b['staff']} {cur}{b['total_price']}")
        text="\n\n".join(lines)
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("← Back",callback_data="back")]])
    if edit: await msg.edit_text(text,parse_mode="Markdown",reply_markup=kb)
    else: await msg.reply_text(text,parse_mode="Markdown",reply_markup=kb)

async def _show_revenue(msg,edit=False):
    r=get_revenue_today(); cur=get_setting("currency","₹")
    with get_db() as db:
        week=db.execute("SELECT COALESCE(SUM(advance_amount),0) as r FROM bookings WHERE date>=date('now','-7 days') AND payment_status='paid'").fetchone()["r"]
        month=db.execute("SELECT COALESCE(SUM(advance_amount),0) as r FROM bookings WHERE date>=date('now','start of month') AND payment_status='paid'").fetchone()["r"]
        total=db.execute("SELECT COALESCE(SUM(advance_amount),0) as r FROM bookings WHERE payment_status='paid'").fetchone()["r"]
    text=f"💰 *Revenue*\n\nToday: {cur}{r['revenue']:,}\nThis Week: {cur}{week:,}\nThis Month: {cur}{month:,}\nAll Time: {cur}{total:,}"
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("← Back",callback_data="back")]])
    if edit: await msg.edit_text(text,parse_mode="Markdown",reply_markup=kb)
    else: await msg.reply_text(text,parse_mode="Markdown",reply_markup=kb)

async def _show_clients(msg,edit=False):
    clients=get_recent_clients(10)
    lines=["👥 *Recent Clients*\n"]
    for c in clients: lines.append(f"👤 *{c['name']}*\n   📞 {c['phone'] or '—'} | 🗓 {c['visit_count']} visits")
    text="\n\n".join(lines) if clients else "No clients yet."
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("← Back",callback_data="back")]])
    if edit: await msg.edit_text(text,parse_mode="Markdown",reply_markup=kb)
    else: await msg.reply_text(text,parse_mode="Markdown",reply_markup=kb)

async def _send_export(msg,context):
    try:
        import openpyxl; from openpyxl.styles import Font,PatternFill
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Bookings"
        headers=["ID","Client","Phone","Service","Staff","Date","Slot","Total","Advance","Payment","Conflict"]
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=1,column=col,value=h)
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill(start_color="D4A017",end_color="D4A017",fill_type="solid")
        with get_db() as db:
            rows=db.execute("SELECT * FROM bookings ORDER BY date DESC").fetchall()
        for i,b in enumerate(rows,2):
            ws.append([b["id"],b["client_name"],b["phone"],b["service"],b["staff"],
                       b["date"],b["slot"],b["total_price"],b["advance_amount"],
                       b["payment_status"],"⚠️" if b["conflict_flag"] else "OK"])
        path="/tmp/salon_report.xlsx"; wb.save(path)
        with open(path,"rb") as f:
            await context.bot.send_document(chat_id=ADMIN_CHAT_ID,document=f,filename="salon_report.xlsx",caption="📊 Salon Report")
    except Exception as e: await msg.reply_text(f"❌ Export failed: {e}")

async def _show_my_bookings(update,context):
    uid=update.effective_user.id; cur=get_setting("currency","₹")
    with get_db() as db:
        rows=db.execute("SELECT * FROM bookings WHERE telegram_id=? ORDER BY date DESC LIMIT 5",(uid,)).fetchall()
    if not rows: text="No bookings yet. Use /book! ✦"
    else:
        lines=["✦ *Your Bookings*\n"]
        for b in rows: lines.append(f"{'✅' if b['status']=='confirmed' else '❌'} `{b['id']}`\n🪄 {b['service']} | 📅 {b['date']} ⏰ {b['slot']}\n💰 {cur}{b['total_price']}")
        text="\n\n".join(lines)
    await update.callback_query.edit_message_text(text,parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✦ Book Again",web_app=WebAppInfo(url=MINI_APP_URL))]]))

async def _show_about(msg,edit=False):
    text=f"✦ *{get_setting('salon_name')}*\n\n📍 {get_setting('salon_address')}\n📞 {get_setting('salon_phone')}\n🕐 {get_setting('open_time')} – {get_setting('close_time')}"
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("✦ Book Now",web_app=WebAppInfo(url=MINI_APP_URL))]])
    if edit: await msg.edit_text(text,parse_mode="Markdown",reply_markup=kb)
    else: await msg.reply_text(text,parse_mode="Markdown",reply_markup=kb)

# ── Reminders ─────────────────────────────────────────────────────────
async def send_reminders(bot):
    now=datetime.now()
    with get_db() as db:
        upcoming=db.execute("SELECT * FROM bookings WHERE status='confirmed' AND reminder_sent<2 AND date>=date('now')").fetchall()
    for b in upcoming:
        try:
            slot_dt=datetime.strptime(f"{b['date']} {b['slot']}","%Y-%m-%d %H:%M")
            hours=(slot_dt-now).total_seconds()/3600; tid=b["telegram_id"]
            if not tid or tid==0: continue
            if b["reminder_sent"]<1 and 23<=hours<=25:
                await bot.send_message(chat_id=tid,parse_mode="Markdown",
                    text=f"🔔 *Reminder!*\n\nAppointment tomorrow!\n🪄 {b['service']} with {b['staff']}\n📅 {b['date']} ⏰ {b['slot']}\n🔖 `{b['id']}`")
                with get_db() as db2: db2.execute("UPDATE bookings SET reminder_sent=1 WHERE id=?",(b["id"],))
            elif b["reminder_sent"]<2 and 0.75<=hours<=1.25:
                await bot.send_message(chat_id=tid,parse_mode="Markdown",
                    text=f"⏰ *In 1 hour!*\n\n🪄 {b['service']} with {b['staff']} at {b['slot']}\n📍 {get_setting('salon_address')}\n_See you soon! 🌸_")
                with get_db() as db2: db2.execute("UPDATE bookings SET reminder_sent=2 WHERE id=?",(b["id"],))
        except Exception as e: logger.warning(f"[REMINDER] {b['id']}: {e}")

# ── Main ──────────────────────────────────────────────────────────────
def run_backend():
    uvicorn.run(api,host="0.0.0.0",port=PORT,log_level="warning")

async def run_bot():
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    tg_app=Application.builder().token(BOT_TOKEN).build()
    tg_app.add_handler(CommandHandler("start",start))
    tg_app.add_handler(CommandHandler("book",start))
    tg_app.add_handler(CommandHandler("admin",admin_cmd))
    tg_app.add_handler(CommandHandler("bookings_today",lambda u,c: _show_today(u.message) if is_admin(u.effective_user.id) else None))
    tg_app.add_handler(CommandHandler("revenue_today",lambda u,c: _show_revenue(u.message) if is_admin(u.effective_user.id) else None))
    tg_app.add_handler(CommandHandler("clients",lambda u,c: _show_clients(u.message) if is_admin(u.effective_user.id) else None))
    tg_app.add_handler(CommandHandler("export",lambda u,c: _send_export(u.message,c) if is_admin(u.effective_user.id) else None))
    tg_app.add_handler(CommandHandler("status",status_cmd))
    tg_app.add_handler(CommandHandler("setname",setname_cmd))
    tg_app.add_handler(CommandHandler("sethours",sethours_cmd))
    tg_app.add_handler(CommandHandler("addservice",addservice_cmd))
    tg_app.add_handler(CommandHandler("addstaff",addstaff_cmd))
    tg_app.add_handler(CallbackQueryHandler(handle_callback))
    scheduler=AsyncIOScheduler()
    scheduler.add_job(lambda:asyncio.create_task(send_reminders(tg_app.bot)),"interval",minutes=15)
    scheduler.start()
    await tg_app.initialize()
    await tg_app.start()
    await tg_app.updater.start_polling(allowed_updates=Update.ALL_TYPES)
    logger.info(f"✦ Salon Bot v3 live | Admin: {ADMIN_CHAT_ID}")
    try: await asyncio.Event().wait()
    finally:
        await tg_app.updater.stop(); await tg_app.stop(); await tg_app.shutdown()

def main():
    init_db()
    threading.Thread(target=run_backend,daemon=True).start()
    logger.info(f"[BACKEND] Started on port {PORT}")
    asyncio.run(run_bot())

if __name__=="__main__":
    main()
